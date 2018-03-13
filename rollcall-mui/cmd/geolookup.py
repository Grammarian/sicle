from openpyxl import load_workbook
import os.path

import json
import requests

SOURCE_XLSX = "./Schools Locations Classes - Feb 2018.xlsx"
TITLES = ["Organisation", "School", "Location", "IBN ID", "File No", "Language", "Classes",
          "Address1", "Address2", "Suburb", "State", "Postcode", "Contact Name", "Contact Phone", "Electorate", "Education Officer"]


def path_to_xlsx():
    path_to_py = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(path_to_py, SOURCE_XLSX)


def process_row(row):
    location = row[2].value
    # print(tuple(x.value for x in row))
    print(location)

    payload = {
        "key": "AIzaSyDQkcz38z2AbEYTuSIXDj66DEixa7LCDhs",
        "query": location,
        "type": "school"
    }
    response = requests.get(
        "https://maps.googleapis.com/maps/api/place/textsearch/json", params=payload)
    json_data = json.loads(response.text)
    if json_data["status"] == "OK":
        result1 = json_data["results"][0]
        data = {
            "name": result1["name"],
            "found": True,
            "lat": result1["geometry"]["location"]["lat"],
            "lng": result1["geometry"]["location"]["lng"]
        }
        print(data)
    else:
        data = {
            "name": result1["name"],
            "found": False
        }
        print(data)


def is_title_row(row):
    return all(x.value == TITLES[i] for (i, x) in enumerate(row))


def main():
    wb = load_workbook(filename=path_to_xlsx(), data_only=False)
    sheet = wb["Sheet0"]
    found_titles = False
    for row in sheet.iter_rows():
        if found_titles:
            process_row(row)
        else:
            found_titles = is_title_row(row)


if __name__ == "__main__":
    main()
