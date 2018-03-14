# pip install openpyxl
# pip install cuid
import os.path
import json
import datetime

from openpyxl import load_workbook
import cuid  # https://github.com/necaris/cuid.py - create uuid's in the format that graphcool expects


SOURCE_XLSX = "./data/CLP-DATAEXTRACT.xlsx"
SCHOOL_TITLES = ["ORGANISATION_ID", "ORGANISATION_NAME", "ORG_ELECTORATE", "P_ADDRESS1", "P_SUBURB", "P_STATE",
                 "P_POSTCODE", "S_ADDRESS1", "S_SUBURB", "S_STATE", "S_POSTCODE", "SCHOOL_NAME", "SCH_ELECTORATE",
                 "SCHOOL_ID", "SCHOOL_P_ADDRESS1",
                 "SCHOOL_P_SUBURB", "SCHOOL_P_STATE", "SCHOOL_P_POSTCODE", "SCHOOL_S_ADDRESS1", "SCHOOL_S_SUBURB",
                 "SCHOOL_S_STATE", "SCHOOL_S_POSTCODE", "LOCATION_NAME", "LOC_ELECTORATE", "LOC_S_ADDRESS1",
                 "LOC_S_SUBURB", "LOC_S_STATE", "LOC_S_POSTCODE"]
ORGANISATION_FIELDS = {"ORGANISATION_ID": "CLP_ORGANISATION_ID", "ORGANISATION_NAME": "NAME",
                       "ORG_ELECTORATE": "ELECTORATE", "S_ADDRESS1": "ADDRESS", "S_SUBURB": "SUBURB",
                       "S_STATE": "STATE", "S_POSTCODE": "POSTCODE", }
SCHOOL_FIELDS = {"SCHOOL_NAME": "NAME", "SCH_ELECTORATE": "ELECTORATE", "SCHOOL_ID": "CLP_SCHOOL_ID",
                 "ORGANISATION_ID": "CLP_ORGANISATION_ID",
                 "SCHOOL_S_ADDRESS1": "ADDRESS", "SCHOOL_S_SUBURB": "SUBURB", "SCHOOL_S_STATE": "STATE",
                 "SCHOOL_S_POSTCODE": "POSTCODE", }
LOCATION_FIELDS = {"LOCATION_NAME": "NAME", "LOC_ELECTORATE": "ELECTORATE", "SCHOOL_ID": "CLP_SCHOOL_ID",
                   "LOC_S_ADDRESS1": "ADDRESS", "LOC_S_SUBURB": "SUBURB", "LOC_S_STATE": "STATE",
                   "LOC_S_POSTCODE": "POSTCODE"}
TEACHER_TITLES = ["TEACHER_ID", "ORGANISATION_NAME", "SCHOOL_NAME", "TEACHER_NAME", "LNAME", "FNAME",
                  "TEACHER_LANGUAGES", "P_ADDRESS1", "P_ADDRESS2", "P_SUBURB", "P_STATE", "P_POSTCODE",
                  "ORGANISATION_ID", "SCHOOL_ID"]
STUDENT_TITLES = ["SCHOOL_NAME", "SCHOOL_ID", "STUDENT_ID", "LOCATION_NAME",
                  "STUDENT_LNAME", "STUDENT_FNAME", "DOB", "TEL", "LOCATION_NAME_1"]
TEACHER_FIELDS = {"TEACHER_ID": "CLP_TEACHER_ID", "ORGANISATION_NAME": "ORGANISATION_NAME",
                  "SCHOOL_NAME": "SCHOOL_NAME",
                  "LNAME": "FAMILY_NAME", "FNAME": "GIVEN_NAMES", "TEACHER_LANGUAGES": "LANGUAGES",
                  "ORGANISATION_ID": "ORGANISATION_ID", "SCHOOL_ID": "SCHOOL_ID", }
STUDENT_FIELDS = {"SCHOOL_NAME": "SCHOOL_NAME", "SCHOOL_ID": "SCHOOL_ID", "STUDENT_ID": "CLP_STUDENT_ID",
                  "LOCATION_NAME": "LOCATION",
                  "STUDENT_LNAME": "FAMILY_NAME", "STUDENT_FNAME": "GIVEN_NAMES", "DOB": "DATE_OF_BIRTH",
                  "TEL": "PHONE", "LOCATION_NAME_1": "DAY_SCHOOL", }


class Sheet:
    "Data container object to hold the contents of one sheet within an excel spreadsheet"

    def __init__(self, name, titles=None, rows=None):
        self.name = name
        self.titles = titles or []
        self.rows = rows or []


def convert_row_to_dict(titles, row):
    data = {}
    for (i, cell) in enumerate(row):
        data[titles[i]] = str(cell.value)
    return data


def convert_xlsx(xlsx_file):
    """Convert the given XLSX spreadsheet to iterable of Sheet objects,
    in which row has been converted into a dictionary"""
    work_book = load_workbook(filename=xlsx_file, read_only=True, data_only=True)
    for sheet in work_book:
        rows = [x for x in sheet.iter_rows()]
        if rows:
            titles = [cell.value for cell in rows[0]]
            dicts = [convert_row_to_dict(titles, row) for row in rows[1:]]
            yield Sheet(sheet.title, titles, dicts)
        else:
            yield Sheet(sheet.title)


def to_camel(s):
    """Convert an underscored title into camel case. 'PARENT_ORGANISATION_ID' => 'parentOrganisationId'"""
    bits = [(x.lower() if i == 0 else x.title())
            for (i, x) in enumerate(s.split("_"))]
    return "".join(bits)


def relative_to_absolute(relative_path):
    path_to_py = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(path_to_py, relative_path)


def extract(fields, row_as_dict):
    data = {}
    for (k, v) in fields.items():
        data[to_camel(v)] = row_as_dict[k]
    return data


def process_sheet(sheet, titles, field_defns):
    if titles != sheet.titles:
        print("Sheet doesn't have expected titles")
        return []

    structs = [[extract(defn, x) for x in sheet.rows] for defn in field_defns]
    return structs


def unique(key, dicts):
    t = {x[key]: x for x in dicts}
    return t.values()


def now_as_iso8601():
    return datetime.datetime.now().replace(microsecond=0).isoformat() + "Z"


def inject_required(type_name, dicts):
    "Inject the required fields that graphcool import required"
    for x in dicts:
        x["_typeName"] = type_name
        x["id"] = cuid.cuid()
        x["createdAt"] = x["updatedAt"] = now_as_iso8601()
    return dicts


def prepare_organisations(organisations):
    unique_orgs = unique("clpOrganisationId", organisations)
    fat_orgs = inject_required("ClpOrganisation", unique_orgs)
    return fat_orgs


def prepare_schools(schools):
    uniques = unique("clpSchoolId", schools)
    injected = inject_required("ClpSchool", uniques)
    return injected


def prepare_locations(locations):
    # There are multiple locations, each of which is identitical except that for being related to a different school.
    # We have to collect all the schools that meet at the same location.
    uniques = {}
    for x in locations:
        # get an existing location with the given name, or add the new location
        location = uniques.setdefault(x["name"], x)
        related_schools = location.setdefault("schools", list())
        related_schools.append(x.pop("clpSchoolId"))
    injected = inject_required("ClpLocation", uniques.values())

    # FIX THIS - Current extract doesn't include the CLP location id :( Make one up for the time being
    for x in injected:
        x["clpLocationId"] = cuid.cuid()
    return injected


def convert_dob_to_datetime(s):
    "Convert the string from 99/MON/YY to a ISO date"
    dt = datetime.datetime.strptime(s, "%d/%b/%y")
    return dt.isoformat() + ".0Z"  # GraphCool import insists on microseconds, hence the ".0"


def prepare_students(students):
    uniques = unique("clpStudentId", students)
    injected = inject_required("ClpStudent", uniques)
    for x in injected:
        x["dateOfBirth"] = convert_dob_to_datetime(x["dateOfBirth"])
    return injected


def prepare_teachers(teachers):
    # Like locations, the same teacher can have multiple records,
    # each of which is identitical except that for being related to a different school.
    # We have to collect all the schools that the same teacher is teaching at.
    uniques = {}
    for x in teachers:
        # get an existing teacher with that id, or add the new teacher record
        teacher = uniques.setdefault(x["clpTeacherId"], x)
        related_schools = teacher.setdefault("schools", list())
        related_schools.append(x.pop("schoolId"))
    injected = inject_required("ClpTeacher", uniques.values())
    return injected


def extract_from_xlsx(file_path):
    for sheet in convert_xlsx(file_path):
        if sheet.name == "SCHOOL-ORG":
            (organisations, schools, locations) = process_sheet(
                sheet, SCHOOL_TITLES, [ORGANISATION_FIELDS, SCHOOL_FIELDS, LOCATION_FIELDS])
        elif sheet.name == "TEACHER":
            (teachers, ) = process_sheet(sheet, TEACHER_TITLES, [TEACHER_FIELDS])
        elif sheet.name == "STUDENT":
            (students, ) = process_sheet(sheet, STUDENT_TITLES, [STUDENT_FIELDS])
        else:
            print("Ignoring sheet:", sheet.name)
    return (organisations, schools, locations, teachers, students)


def copy_without(dicts, *keys_to_remove):
    "Return iterable that contains copies of the given dictionary with all the given keys removed"
    copies = [x.copy() for x in dicts]
    for d in copies:
        for to_remove in keys_to_remove:
            d.pop(to_remove, None)
    return copies


EXTRACT_OUTPUT_DIR = "../../server/extract"


def write_nodes(file_name, collection):
    path = relative_to_absolute(os.path.join(EXTRACT_OUTPUT_DIR, "nodes", file_name))
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        nodes = {
            "valueType": "nodes",
            "values": collection
        }
        f.write(json.dumps(nodes))


def chunks(l, n):
    """Yield n successive similar-sized chunks from l."""
    chunk_size = 1 + len(l) // n
    for i in range(0, len(l), chunk_size):
        yield l[i:i + chunk_size]


def main():
    (raw_organisations, raw_schools, raw_locations, raw_teachers,
     raw_students) = extract_from_xlsx(relative_to_absolute(SOURCE_XLSX))
    organisations = prepare_organisations(raw_organisations)
    schools = prepare_schools(raw_schools)
    locations = prepare_locations(raw_locations)
    teachers = prepare_teachers(raw_teachers)
    students = prepare_students(raw_students)
    # write_nodes("1.json", list(organisations))
    # write_nodes("2.json", copy_without(schools, "clpOrganisationId"))
    # write_nodes("3.json", copy_without(locations, "schools"))
    # write_nodes("4.json", copy_without(teachers, "organisationId", "organisationName", "schools", "schoolName"))
    expegated_students = copy_without(students, "schoolId", "schoolName", "location")
    for (i, chunk) in enumerate(chunks(expegated_students, 5)):
        write_nodes("%d.json" % (1 + i), chunk)


if __name__ == "__main__":
    main()
